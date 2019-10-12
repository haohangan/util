from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import sys


# 处理数据
workbook = load_workbook(filename="demo.xlsx")
ws = workbook[workbook.sheetnames[0]]
ws.merge_cells('B2:B8') ##合并
ws['B2'].alignment = Alignment(horizontal="center", vertical="center") ##居中

#保存为新文件
newfilename = "demo_new.xlsx"
workbook.save(filename=newfilename)
